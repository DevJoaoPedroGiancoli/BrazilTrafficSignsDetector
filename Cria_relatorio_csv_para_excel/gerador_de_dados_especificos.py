from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from dados_especificos_trecho import *


# Carregar o arquivo Excel existente
existing_wb = load_workbook('Alimentador.xlsx')

# Obter a guia "Especificos_trecho" ou criar uma nova se não existir
ws = existing_wb.get_sheet_by_name('Especificos_trecho') if 'Especificos_trecho' in existing_wb.sheetnames else existing_wb.create_sheet('Especificos_trecho')

#Alimentando dados do trecho de placas em específico

cell_a1 = ws.cell(row=1,column=1)
cell_a1.value = "Sentido:"

cell_a2 = ws.cell(row=1, column=2)
cell_a2.value = sentido_especifico

cell_b1 = ws.cell(row=2, column=1)
cell_b1.value = "Rota:"

cell_b2 = ws.cell(row=2, column=2)
cell_b2.value = rota

cell_c1 = ws.cell(row=3, column=1)
cell_c1.value = "Km_inicial:"

cell_c2 = ws.cell(row=3, column=2)
cell_c2.value = km_inicial

cell_d1 = ws.cell(row=4, column=1)
cell_d1.value = "Km_final:"

cell_d2 = ws.cell(row=4, column=2)
cell_d2.value = km_final

cell_e1 = ws.cell(row=5, column=1)
cell_e1 = "Classe_de_placas_total"

cell_g1= ws.cell(row=6, column=1)
cell_g1.value = "Intervalo de quilometragem:"

# Obter dados de placas

for index, classe in enumerate(nomenclatura_escpecifica.values(), start = 2):
    ws.cell(row=5, column=index, value=classe)

# Objter intervalo de quilometragem

for index, km in enumerate(intervalo_km_dict.values(), start=2):
    ws.cell(row=6, column=index, value=km)

# Salvar a planilha em um arquivo Excel

existing_wb.save('Alimentador.xlsx')