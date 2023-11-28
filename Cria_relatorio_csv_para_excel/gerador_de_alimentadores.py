from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from dados_especificos_trecho import *


# Carregar o arquivo Excel existente
existing_wb = load_workbook('Alimentador.xlsx')

# Obter a guia "Quantidade_km"
ws = existing_wb.get_sheet_by_name('Quantidade_km') if 'Quantidade_km' in existing_wb.sheetnames else existing_wb.create_sheet('Quantidade_km')

# Obter dados de placas

cell_b1 = ws.cell(row=2, column=2, value="Rota:")
cell_c2 = ws.cell(row=2, column=3)
cell_c2.value = rota
cell_b3 = ws.cell(row=3, column=2, value="Km_inicial:")
cell_c3 = ws.cell(row=3, column=3)
cell_c3.value = km_inicial
cell_b4 = ws.cell(row=4, column=2, value="Km_final:")
cell_c4= ws.cell(row=4, column=3)
cell_c4.value = km_final
cell_d2 = ws.cell(row=2, column=4, value="Sentido:")
cell_e2 = ws.cell(row=2, column=5)
cell_e2.value = sentido_especifico

g = 7

cell_g2 = ws.cell(row=2, column=7, value="Quantidade (KM/Placa)")

for index, classe in enumerate(nomenclatura_escpecifica.values(), start = 8):
    ws.cell(row=2, column=index, value=classe)


for index, km in enumerate(intervalo_km_dict.values(), start=3):
    ws.cell(row=index, column=7, value=km)

# Obter a guia "Densidade_Area"
ws = existing_wb.get_sheet_by_name('Densidade_Area') if 'Densidade_Area' in existing_wb.sheetnames else existing_wb.create_sheet('Densidade_Area')

cell_g2 = ws.cell(row=2, column=7, value="Densidade (KM/Placa)")


cell_b1 = ws.cell(row=2, column=2, value="Rota:")
cell_c2 = ws.cell(row=2, column=3)
cell_c2.value = rota
cell_b3 = ws.cell(row=3, column=2, value="Km_inicial:")
cell_c3 = ws.cell(row=3, column=3)
cell_c3.value = km_inicial
cell_b4 = ws.cell(row=4, column=2, value="Km_final:")
cell_c4= ws.cell(row=4, column=3)
cell_c4.value = km_final
cell_d2 = ws.cell(row=2, column=4, value="Sentido:")
cell_e2 = ws.cell(row=2, column=5)
cell_e2.value = sentido_especifico

for index, classe in enumerate(nomenclatura_escpecifica.values(), start = 8):
    ws.cell(row=2, column=index, value=classe)


for index, km in enumerate(intervalo_km_dict.values(), start=3):
    ws.cell(row=index, column=7, value=km)

# Obter a guia "Area"
ws = existing_wb.get_sheet_by_name('Area') if 'Area' in existing_wb.sheetnames else existing_wb.create_sheet('Area')

cell_b1 = ws.cell(row=2, column=2, value="Rota:")
cell_c2 = ws.cell(row=2, column=3)
cell_c2.value = rota
cell_b3 = ws.cell(row=3, column=2, value="Km_inicial:")
cell_c3 = ws.cell(row=3, column=3)
cell_c3.value = km_inicial
cell_b4 = ws.cell(row=4, column=2, value="Km_final:")
cell_c4= ws.cell(row=4, column=3)
cell_c4.value = km_final
cell_d2 = ws.cell(row=2, column=4, value="Sentido:")
cell_e2 = ws.cell(row=2, column=5)
cell_e2.value = sentido_especifico

cell_g2 = ws.cell(row=2, column=7, value="Area total por placa")


for index, classe in enumerate(nomenclatura_escpecifica.values(), start = 3):
    ws.cell(row=index, column=7, value=classe)
      
print("Alimentadores com bem-sucedidos com a alimentação dos dados..")
existing_wb.save('Alimentador.xlsx')