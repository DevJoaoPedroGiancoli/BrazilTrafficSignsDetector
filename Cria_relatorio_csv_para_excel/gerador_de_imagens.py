from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from dados_especificos_trecho import *
from gerador_de_dados_especificos import *
from gerador_de_dados_brutos import *
from gerador_de_alimentadores import *
from gerador_de_dados_tabela import *

# Carregar o arquivo Excel existente
existing_wb = load_workbook('Alimentador.xlsx')

# Obter a guia "Imagens_cortadas"
ws = existing_wb.get_sheet_by_name('Imagens_Cortadas') if 'Imagens_Cortadas' in existing_wb.sheetnames else existing_wb.create_sheet('Imagens_Cortadas')

cell_b2 = ws.cell(row=3,column=2, value="ID")
cell_c2 = ws.cell(row=3,column=3, value="Placa")
cell_d2 = ws.cell(row=3,column=4, value="Crop")

# Configurar a largura da coluna "C"
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 15

# Itere sobre as linhas do DataFrame
for index, row in merged_data.iterrows():
    
    ws.row_dimensions[index+4].height = 40
    
    # Classe da Placa
    cell_c = ws.cell(row=index+4,column=3)
    sign_class_id = row['Sign_class']
    sign_class_value = sign_class_dict.get(sign_class_id,'')
    cell_c.value = sign_class_value
    cell_c.border = Border(top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    # Identificador único por placa
    cell_b = ws.cell(row=index+4, column=2)
    cell_b.value = row['ID']
    cell_b.border = Border(top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    # Inserir a imagem recortadas correspondentes aos IDs
    img_id = int(row['ID'])
    img_pattern = f"imagem_destacada_do_id_{img_id}_*"
    img_files = glob.glob(os.path.join(imagens_folder, img_pattern + ".jpg"))

    img_file = max(img_files, key=os.path.getctime)  # Use o arquivo mais recente
    img_destacada = Image(img_file)
    cell_d = ws.cell(row=index+4, column=4)  # Coluna 'Imagem Recortada'
    ws.add_image(img_destacada, cell_d.coordinate)  # Insere a imagem na célula D correspondente
    
    # Configurar o tamanho da imagem
    img_destacada.width = 1.5 * 28.35  # 0.95 cm convertido para pixels (1 cm = 28.35 pixels)
    img_destacada.height = 1.5 * 28.35  # 0.65 cm convertido para pixels
    
    cell_d.alignment = Alignment(horizontal='center', vertical='center')
    cell_d.border = Border(right=Side(style='dashed'), top=Side(style='dashed'), bottom=Side(style='dashed'))
    


print("Tabelas de imagem cortadas salva..")


# Obter a guia "Quantidade_km"
ws = existing_wb.get_sheet_by_name('Imagens_Inteiras') if 'Imagens_Inteiras' in existing_wb.sheetnames else existing_wb.create_sheet('Imagens_Inteiras')

cell_b3 = ws.cell(row=3,column=2, value="ID")
cell_c3 = ws.cell(row=3,column=3, value="Placa")
cell_d4 = ws.cell(row=3,column=4, value="Imagem_Inteira")

# Configurar a largura da coluna "C"
ws.column_dimensions['B'].width = 8.10
ws.column_dimensions['C'].width = 8.10
ws.column_dimensions['D'].width = 58

# Itere sobre as linhas do DataFrame
for index, row in merged_data.iterrows():
    
    ws.row_dimensions[index+4].height = 261
    
    # Classe da Placa
    cell_c = ws.cell(row=index+4,column=3)
    sign_class_id = row['Sign_class']
    sign_class_value = sign_class_dict.get(sign_class_id,'')
    cell_c.value = sign_class_value
    cell_c.border = Border(top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    # Identificador único por placa
    cell_b = ws.cell(row=index+4, column=2)
    cell_b.value = row['ID']
    cell_b.border = Border(top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    # Inserir a imagem recortadas correspondentes aos IDs
    img_id = int(row['ID'])
    img_pattern = f"foto_inteira_do_id_{img_id}_*"
    img_files = glob.glob(os.path.join(imagens_folder_inteiras, img_pattern + ".jpg"))

    img_file = max(img_files, key=os.path.getctime)  # Use o arquivo mais recente
    img_destacada = Image(img_file)
    cell_d = ws.cell(row=index+4, column=4)  # Coluna 'Imagem Recortada'
    ws.add_image(img_destacada, cell_d.coordinate)  # Insere a imagem na célula D correspondente
    
    # Configurar o tamanho da imagem
    img_destacada.width = 13 * 28.35  # 0.95 cm convertido para pixels (1 cm = 28.35 pixels)
    img_destacada.height = 10.58 * 28.35  # 0.65 cm convertido para pixels
    
    cell_d.alignment = Alignment(horizontal='center', vertical='center')
    cell_d.border = Border(right=Side(style='dashed'), top=Side(style='dashed'), bottom=Side(style='dashed'))
    
    

print("Imagens inteiras salvas..")
existing_wb.save('Alimentador.xlsx')

print("Relatorio salvo com sucesso!!!!!")