import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

# Carregar os dados do CSV em um DataFrame
data = pd.read_csv(r'save_test\seusassaasarquivo_unico.csv')

# Criar uma nova planilha do Excel
wb = Workbook()
ws = wb.active

# Iterar pelas linhas do DataFrame e inserir os dados e imagens
for index, row in data.iterrows():
    # Inserir os valores nas células B e C
    ws.cell(row=index+2, column=2, value=row['Sign_class'])
    ws.cell(row=index+2, column=3, value=row['ID'])

    # Centralizar os valores nas células B e C
    for col in range(2, 4):  # Colunas B e C
        cell = ws.cell(row=index+2, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Inserir imagens correspondentes aos IDs
    img_path = f"save_test\cropped_img_{int(row['ID'])}.jpg"  # Nome do arquivo de imagem
    img = Image(img_path)

    # Obter as dimensões da imagem
    img_width, img_height = img.width, img.height

    # Definir o tamanho da célula D com base nas dimensões da imagem
    cell_id = f'D{index+2}'
    ws.column_dimensions[cell_id[0]].width = img_width  # Ajusta a largura da célula
    ws.row_dimensions[int(cell_id[1:])].height = img_height  # Ajusta a altura da célula

    # Inserir a imagem na célula D especifica
    cell_d = ws.cell(row=index+2, column=4)
    ws.add_image(img, cell_d.coordinate)  # Insere a imagem na célula D correspondente

    # Centralizar a imagem horizontalmente na célula D
    cell_d.alignment = Alignment(horizontal='center', vertical='center')

# Salvar a planilha em um arquivo Excel
wb.save('planilha_com_imagens_centralizadas_5.xlsx')
