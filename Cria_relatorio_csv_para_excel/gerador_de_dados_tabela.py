import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
from dados_especificos_trecho import *
import glob
import os


# Carregar os dados do CSV em um DataFrame
data_placas = pd.read_csv(r'Resultado_de_dados\arquivos_csv\classe_de_placas_por_id.csv')
data_horarios = pd.read_csv(r'Resultado_de_dados\arquivos_csv\horarios_por_id.csv',  encoding='latin-1')
merged_data = data_horarios.merge(data_placas, on='ID', how='inner')

# Sign_class dicionario:

sign_class_dict = {0: 'A-10b', 1: 'A-11b', 2: 'A-12', 3: 'A-13a', 4: 'A-13b'
                   , 5: 'A-14', 6: 'A-15', 7: 'A-18', 8: 'A-1a', 9: 'A-1b'
                   , 10: 'A-20a', 11: 'A-20b', 12: 'A-21c', 13: 'A-21d'
                   , 14: 'A-21e', 15: 'A-22', 16: 'A-24', 17: 'A-27'
                   , 18: 'A-28', 19: 'A-2a', 20: 'A-2b', 21: 'A-31', 22: 'A-32a'
                   , 23: 'A-32b', 24: 'A-33a', 25: 'A-33b', 26: 'A-3a', 27: 'A-3b'
                   , 28: 'A-42a', 29: 'A-42b', 30: 'A-4a', 31: 'A-4b', 32: 'A-52',
                   33: 'A-5a', 34: 'A-5b', 35: 'A-6', 36: 'A-7a', 37: 'A-7b', 38: 'A-8'
                   , 39: 'Del', 40: 'E-5', 41: 'ESP-20', 42: 'I-4', 43: 'LOC-6', 44: 'MO'
                   , 45: 'MP', 46: 'R-1', 47: 'R-15', 48: 'R-19', 49: 'R-2', 50: 'R-24a',
                   51: 'R-24b', 52: 'R-26', 53: 'R-27', 54: 'R-28', 55: 'R-33', 56: 'R-43',
                   57: 'R-4a', 58: 'R-4b', 59: 'R-5a', 60: 'R-6a', 61: 'R-6b', 62: 'R-6c', 
                   63: 'R-7', 64: 'RQ', 65: 'S-14', 66: 'TUR-4', 67: 'de'}



# Carregar o arquivo Excel existente
existing_wb = load_workbook('Alimentador.xlsx')

# Obter a guia "Dados_Brutos" ou criar uma nova se não existir
ws = existing_wb.get_sheet_by_name('Dados_Brutos') if 'Dados_Brutos' in existing_wb.sheetnames else existing_wb.create_sheet('Dados_Brutos')


# Adicione cabeçalhos às colunas e estilize-os
#              1          2       3            4             5               6                    7
headers = ["Contador","Classe", "ID", "Horario" ,"Coordenadas_por_GPS","Km da Rota","Imagem_cortada", "Imagem_inteira", "Area da Placa" ]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num, value=header)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Adicione filtro ao cabeçalho
ws.auto_filter.ref = ws.dimensions


# Itere sobre as linhas do DataFrame
for index, row in merged_data.iterrows():
    
    # Classe da Placa
    cell_b = ws.cell(row=index+3,column=2)
    sign_class_id = row['Sign_class']
    sign_class_value = sign_class_dict.get(sign_class_id,'')
    cell_b.value = sign_class_value
    
    # Identificador único por placa
    cell_c = ws.cell(row=index+3, column=3)
    cell_c.value = row['ID']
    
   
    # Horários
    cell_d = ws.cell(row=index+3,column=4)
    cell_d.value = row['Horario']
    
    # GPS(necessário linkar com dados de Issaias)
    
    cell_e = ws.cell(row=index+3, column=5)
    cell_e.value = "-" 
    cell_f = ws.cell(row=index+3, column=6)
    cell_f.value = "-"

    cell_g = ws.cell(row=index+3, column=7)  # Coluna 'Imagem Cortada'
    cell_g.value = "Imagem_Cortada"
    
    cell_h = ws.cell(row=index+3, column=8)  # Coluna 'Imagem Inteira'
    cell_h.value = "Imagem_Inteira"
    
print("Dados de tabela salvos..")
existing_wb.save('Alimentador.xlsx')