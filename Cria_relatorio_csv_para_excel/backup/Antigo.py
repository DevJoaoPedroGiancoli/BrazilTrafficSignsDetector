import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import glob
import os

# Tamanho fixo para as imagens cortadas
fixed_image_width = 200
fixed_image_height = 200


# Tamanho fixo para as imagens inteiras
fixed_image_inteirias_width = 400
fixed_image_inteiras_height = 300

# Largura fixa da coluna "Imagem"
image_column_width = 31
# Largura fixa da coluna "Imagem"
image_column_d_width = 58

# Altura fixa das colunas
column_height = 250

# Carregar os dados do CSV em um DataFrame
data_placas = pd.read_csv(r'Resultado_de_dados\arquivos_csv\classe_de_placas_por_id.csv')
data_horarios = pd.read_csv(r'Resultado_de_dados\arquivos_csv\horarios_por_id.csv',  encoding='latin-1')

# Sign_class dicionario:

sign_class_dict = {0: 'A-10b', 1: 'A-11b', 2: 'A-12', 3: 'A-13a', 4: 'A-13b'
                   , 5: 'A-14', 6: 'A-15', 7: 'A-18', 8: 'A-1a', 9: 'A-1b'
                   , 10: 'A-20a', 11: 'A-20b', 12: 'A-21c', 13: 'A-21d'
                   , 14: 'A-21e', 15: 'A-22', 16: 'A-24', 17: 'A-27'
                   , 18: 'A-28', 19: 'A-2a', 20: 'A-2b', 21: 'A-31', 22: 'A-32a'
                   , 23: 'A-32b', 24: 'A-33a', 25: 'A-33b', 26: 'A-3a', 27: 'A-3b'
                   , 28: 'A-42a', 29: 'A-42b', 30: 'A-4a', 31: 'A-4b', 32: 'A-52',
                   33: 'A-5a', 34: 'A-5b', 35: 'A-6', 36: 'A-7a', 37: 'A-7b', 38: 'A-8'
                   , 39: 'Del', 40: 'E-5', 41: 'ESP-20', 42: 'I-4', 43: 'LOC-6', 44: 'MO'
                   , 45: 'MP', 46: 'R-1', 47: 'R-15', 48: 'R-19', 49: 'R-2', 50: 'R-24a',
                   51: 'R-24b', 52: 'R-26', 53: 'R-27', 54: 'R-28', 55: 'R-33', 56: 'R-43',
                   57: 'R-4a', 58: 'R-4b', 59: 'R-5a', 60: 'R-6a', 61: 'R-6b', 62: 'R-6c', 
                   63: 'R-7', 64: 'RQ', 65: 'S-14', 66: 'TUR-4', 67: 'de'}


merged_data = data_horarios.merge(data_placas, on='ID', how='inner')

# Criar uma nova planilha do Excel
wb = Workbook()
ws = wb.active

# Adicione cabeçalhos às colunas e estilize-os
headers = ["Classe", "ID", "Placa_Identificada","Imagem" ,"Horario", "Pista (Simples/Dupla)", "Sentido", "Data da medição"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Adicione filtro ao cabeçalho
ws.auto_filter.ref = ws.dimensions

# Inserir imagens inteiras correspondentes aos IDs

# Pasta onde as imagens estão localizadas
imagens_folder = "Resultado_de_dados/imagens_cortadas/"

# Itere sobre as linhas do DataFrame
for index, row in merged_data.iterrows():
    img_id = int(row['ID'])
    img_pattern = f"imagem_destacada_do_id_{img_id}_*"
    img_files = glob.glob(os.path.join(imagens_folder, img_pattern + ".jpg"))

    if img_files:
        img_file = max(img_files, key=os.path.getctime)  # Use o arquivo mais recente
        img = Image(img_file)

        # Definir o tamanho da imagem para um tamanho fixo
        img.width = fixed_image_inteirias_width
        img.height = fixed_image_inteiras_height

        # Inserir a imagem na célula E específica
        cell_e = ws.cell(row=index + 2, column=4)  # Coluna 'Imagem'
        ws.add_image(img, cell_e.coordinate)  # Insira a imagem na célula E correspondente
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_e.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    else:
        print(f"Imagem completa não encontrada para o ID {img_id}")

# Defina alturas e larguras das células apenas a partir da segunda linha
for row_num in range(2, len(merged_data) + 2):  # +2 para incluir a linha de cabeçalho
    ws.column_dimensions['C'].width = image_column_width
    ws.column_dimensions['D'].width = image_column_d_width
    ws.row_dimensions[row_num].height = column_height
    
imagens_folder_inteiras = "Resultado_de_dados/imagens_inteiras/"

# Inserir imagens recortadas correspondentes aos IDs
for index, row in merged_data.iterrows():
    img_id = int(row['ID'])
    img_pattern = f"foto_inteira_do_id_{img_id}_*"
    img_files = glob.glob(os.path.join(imagens_folder, img_pattern + ".jpg"))

    if img_files:
        img_file = max(img_files, key=os.path.getctime)  # Use o arquivo mais recente
        img = Image(img_file)
    

    # Definir o tamanho da imagem para um tamanho fixo
    img.width = fixed_image_width
    img.height = fixed_image_height

    # Inserir a imagem na célula D especifica
    cell_d = ws.cell(row=index+2, column=3)  # Coluna 'Imagem'
    ws.add_image(img, cell_d.coordinate)  # Insere a imagem na célula D correspondente
    cell_d.alignment = Alignment(horizontal='center', vertical='center')
    cell_d.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    
    

# Salvar a planilha em um arquivo Excel
wb.save('output2.xlsx')
