import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

# Carregar os dados do CSV em um DataFrame
data_placas = pd.read_csv(r'Resultado_de_dados\arquivos_csv\classe_de_placas_por_id.csv')
data_horarios = pd.read_csv(r'Resultado_de_dados\arquivos_csv\horarios_por_id.csv')

# Criar uma nova coluna 'HoraMinSeg' com o formato desejado
data_horarios['HoraMinSeg'] = data_horarios.apply(lambda row: f"{row['Horas']}:{row['Minutos']}:{row['Segundos']}", axis=1)
merged_data = data_horarios.merge(data_placas, on='ID', how='inner')

# Criar uma nova planilha do Excel
wb = Workbook()
ws = wb.active

# Iterar pelas linhas do DataFrame e inserir os dados e imagens
for index, row in merged_data.iterrows():
    # Inserir os valores nas células B e C
    ws.cell(row=index+2, column=2, value=row['Sign_class'])
    ws.cell(row=index+2, column=3, value=row['ID'])
    ws.cell(row=index + 2, column=5, value=row['HoraMinSeg'])

    # Centralizar os valores nas células B e C
    for col in range(2, 4):  # Colunas B e C
        cell = ws.cell(row=index+2, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Inserir imagens correspondentes aos IDs
    img_path = f"Resultado_de_dados\imagens_cortadas\imagem_destacada_do_id_{int(row['ID'])}.jpg"  # Nome do arquivo de imagem
    img = Image(img_path)

    # Obter as dimensões da imagem
    img_width, img_height = img.width, img.height

    # Definir o tamanho da célula D com base nas dimensões da imagem
    cell_id = f'D{index+2}'
    ws.column_dimensions[cell_id[0]].width = img_width  # Ajusta a largura da célula
    ws.row_dimensions[int(cell_id[1:])].height = img_height  # Ajusta a altura da célula

    # Inserir a imagem na célula D especifica
    cell_d = ws.cell(row=index+2, column=4)
    ws.add_image(img, cell_d.coordinate)  # Insere a imagem na célula D correspondente

    # Centralizar a imagem horizontalmente na célula D
    cell_d.alignment = Alignment(horizontal='center', vertical='center')

# Salvar a planilha em um arquivo Excel
wb.save('output.xlsx')
